"""Loader for the 'Quién es Quién' (QeQ) personality roster.

Reads the Excel files under ``QeQ/`` — a curated roster of ~2,000 Argentine,
Uruguayan, Paraguayan and international public figures — and turns them into
entries compatible with ``known_people.KNOWN_PEOPLE``. The dashboard glossary
uses these entries to identify people appearing in news headlines.

Parsed entries are cached as ``data/qeq_people.json`` so the dashboard does not
need to parse the Excel files on every run. The cache is rebuilt automatically
when any source file is newer than the cache.

Entry schema::

    {
      "name": str,              # canonical full name
      "role": str,              # current position / public role
      "country": str,           # argentina | uruguay | paraguay | international
      "bio": str,               # Spanish factual description (≤220 chars)
      "aliases": [str, ...],    # zero or more alternate names / nicknames
      "source": "qeq",
    }
"""

from __future__ import annotations

import json
import os
import re
from pathlib import Path

from name_utils import (
    NAME_CONNECTORS as _NAME_CONNECTORS,
    NON_PERSON_TOKENS as _NON_PERSON_TOKENS,
    looks_like_person_name as _looks_like_person_name,
    strip_accents as _strip_accents,
)

REPO_ROOT = Path(__file__).resolve().parent
QEQ_DIR = REPO_ROOT / "QeQ"
CACHE_PATH = REPO_ROOT / "data" / "qeq_people.json"

# País → country code used by the dashboard glossary
_COUNTRY_MAP = {
    "argentina": "argentina",
    "uruguay": "uruguay",
    "paraguay": "paraguay",
}

# Headers we care about (normalized lowercase, no accents)
_HDR_NAME = {"nombre completo"}
_HDR_COUNTRY = {"pais"}
_HDR_ROLE = {"cargo / rol", "cargo/rol", "rol historico", "rol"}
_HDR_BIO = {"descripcion", "por que sigue en los medios"}
_HDR_ALIAS = {"alias / apodos (scraper)", "alias / apodos", "alias"}
_HDR_SCOPE = {"ambito"}

# Files where the roster is purely informational (plan / meta docs) — skip.
_SKIP_FILE_PATTERNS = ("plan_expansion",)


def _norm_header(h) -> str:
    if h is None:
        return ""
    return _strip_accents(str(h).strip()).replace("  ", " ")


def _infer_country_from_filename(fname: str) -> str:
    """Fallback when a file lacks a País column (e.g. the Historicas file)."""
    up = fname.upper()
    if "ARG" in up:
        return "argentina"
    if "URU" in up:
        return "uruguay"
    if "PAR" in up:
        return "paraguay"
    if "INT" in up:
        return "international"
    return "international"


def _normalize_country(raw: str, fallback: str) -> str:
    if not raw:
        return fallback
    key = _strip_accents(str(raw).strip())
    return _COUNTRY_MAP.get(key, "international")


def _clean_bio(text: str, max_len: int = 220) -> str:
    if not text:
        return ""
    s = re.sub(r"\s+", " ", str(text).strip())
    # Strip internal cross-reference prefixes like "Ver ARG-173 (batch principal)."
    s = re.sub(
        r"^Ver\s+(?:[A-Z]{2,4}[-‑]?\d{1,4}|#\d{1,4})"
        r"(?:\s*\([^)]*\))?\.?\s*",
        "", s,
    ).strip()
    if len(s) > max_len:
        # Cut at last sentence boundary before the limit, else hard cut.
        cut = s[: max_len - 1]
        last_dot = cut.rfind(". ")
        if last_dot > max_len * 0.6:
            return cut[: last_dot + 1].strip()
        return cut.rstrip() + "…"
    return s


def _parse_aliases(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;|/]", str(raw))
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        # Drop parentheticals like "(scraper)"
        p = re.sub(r"\s*\([^)]*\)\s*$", "", p).strip()
        if len(p) >= 3 and p.lower() not in {a.lower() for a in out}:
            out.append(p)
    return out


def _parse_workbook(path: Path) -> list[dict]:
    import openpyxl  # local import so dashboard doesn't pay cost on cache hit

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Header row is row 3 in every file (row 1 = title, row 2 = subtitle).
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        norm = [_norm_header(c) for c in row]
        if any(h in _HDR_NAME for h in norm):
            header_row = i
            headers = norm
            break
    if header_row is None:
        wb.close()
        return []

    def find_col(candidates: set[str]) -> int | None:
        for idx, h in enumerate(headers):
            if h in candidates:
                return idx
        return None

    col_name = find_col(_HDR_NAME)
    col_country = find_col(_HDR_COUNTRY)
    col_role = find_col(_HDR_ROLE)
    col_bio = find_col(_HDR_BIO)
    col_alias = find_col(_HDR_ALIAS)
    col_scope = find_col(_HDR_SCOPE)

    if col_name is None:
        wb.close()
        return []

    fallback_country = _infer_country_from_filename(path.name)
    out: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[col_name] if col_name < len(row) else None
        if not name or not str(name).strip():
            continue
        name = re.sub(r"\s+", " ", str(name).strip())
        # Handle inverted parenthetical form "Orsi (Yamandú)" → "Yamandú Orsi"
        m = re.match(r"^([A-ZÁÉÍÓÚÑ][\w]+)\s*\(([A-ZÁÉÍÓÚÑ][\w\s]+)\)$", name)
        if m:
            name = f"{m.group(2).strip()} {m.group(1).strip()}"
        # Strip any remaining trailing parentheticals like "(en medios)" so
        # "Yamandú Orsi" and "Yamandú Orsi (en medios)" dedupe to one entry.
        name = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()
        # Skip pseudo-entries that are actually event descriptors, not people
        # ("Milei y el Papa Francisco", "Javier Milei en Davos", etc.)
        if re.search(r"\b(y el|y la|visita|gira|agenda|caso|en el|en la|en los)\b", name, re.I):
            continue
        # Also skip "Name en Place" patterns like "Javier Milei en Davos"
        if re.search(r"\ben\s+[A-Z]", name):
            continue
        # Drop single-word entries — too ambiguous to match reliably
        if len(name.split()) < 2:
            continue
        if not _looks_like_person_name(name):
            continue

        country = _normalize_country(
            row[col_country] if col_country is not None and col_country < len(row) else None,
            fallback_country,
        )

        role = ""
        if col_role is not None and col_role < len(row) and row[col_role]:
            role = re.sub(r"\s+", " ", str(row[col_role]).strip())
        # Scope (Ámbito) as a fallback when role is empty
        if not role and col_scope is not None and col_scope < len(row) and row[col_scope]:
            role = re.sub(r"\s+", " ", str(row[col_scope]).strip())

        bio = ""
        if col_bio is not None and col_bio < len(row) and row[col_bio]:
            bio = _clean_bio(row[col_bio])

        aliases = []
        if col_alias is not None and col_alias < len(row) and row[col_alias]:
            aliases = _parse_aliases(row[col_alias])

        if not role:
            continue  # an unroled entry has nothing to contribute

        out.append({
            "name": name,
            "role": role,
            "country": country,
            "bio": bio,
            "aliases": aliases,
            "source": "qeq",
        })
    wb.close()
    return out


def _collect_source_files() -> list[Path]:
    if not QEQ_DIR.is_dir():
        return []
    files = []
    for p in sorted(QEQ_DIR.iterdir()):
        if p.suffix.lower() != ".xlsx":
            continue
        if any(pat in p.name.lower() for pat in _SKIP_FILE_PATTERNS):
            continue
        files.append(p)
    return files


def _fl_key(name: str) -> str:
    """First+last word key for fuzzy dedup (catches 'Javier Milei' vs 'Javier Gerardo Milei')."""
    words = _strip_accents(name).split()
    return f"{words[0]} {words[-1]}" if len(words) >= 2 else name.lower()


def _merge_into(existing: dict, entry: dict):
    """Merge a new entry's data into an existing one (aliases, bio, role)."""
    merged = list(existing.get("aliases") or [])
    seen_lower = {a.lower() for a in merged}
    for alias in entry.get("aliases") or []:
        if alias.lower() not in seen_lower:
            merged.append(alias)
            seen_lower.add(alias.lower())
    existing["aliases"] = merged
    if len(entry.get("bio", "")) > len(existing.get("bio", "")):
        existing["bio"] = entry["bio"]
    if len(entry.get("role", "")) > len(existing.get("role", "")):
        existing["role"] = entry["role"]


def build_cache() -> list[dict]:
    """Parse every Excel file under QeQ/ and write the JSON cache.

    When the same person appears in multiple source files we merge the
    information: aliases get unioned, and the longest bio wins. This matters
    because the older ``Personalidades_ARG_*`` files lack an alias column,
    while the ``Batch001/002`` files carry nicknames the scraper relies on.

    Deduplication uses both exact name and first+last word matching
    (e.g. 'Javier Milei' and 'Javier Gerardo Milei' merge into one entry,
    keeping the version with the richer aliases/bio).
    """
    # Country preference: when the same person appears with different
    # country tags (e.g. "argentina" and "international"), prefer the
    # Cono-Sur-specific country over "international".
    _COUNTRY_PRIO = {"argentina": 3, "uruguay": 3, "paraguay": 3, "international": 1}

    entries: list[dict] = []
    idx_exact: dict[tuple, int] = {}
    idx_fl: dict[tuple, int] = {}
    # Cross-country FL index: catches duplicates across country boundaries
    # (e.g. Papa Francisco appearing as "argentina" and "international").
    idx_fl_global: dict[str, int] = {}
    for path in _collect_source_files():
        for entry in _parse_workbook(path):
            key_exact = (entry["name"].lower(), entry["country"])
            fl = _fl_key(entry["name"])
            key_fl = (fl, entry["country"])

            # Try exact match first
            if key_exact in idx_exact:
                _merge_into(entries[idx_exact[key_exact]], entry)
                continue

            # Try first+last fuzzy match (same country)
            if key_fl in idx_fl:
                _merge_into(entries[idx_fl[key_fl]], entry)
                continue

            # Try cross-country FL match — merge into existing entry,
            # keeping the more specific country tag.
            if fl in idx_fl_global:
                existing = entries[idx_fl_global[fl]]
                _merge_into(existing, entry)
                # Prefer the more specific country
                if _COUNTRY_PRIO.get(entry["country"], 0) > _COUNTRY_PRIO.get(existing["country"], 0):
                    existing["country"] = entry["country"]
                continue

            idx_exact[key_exact] = len(entries)
            idx_fl[key_fl] = len(entries)
            idx_fl_global[fl] = len(entries)
            entries.append(entry)
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return entries


def _cache_is_fresh() -> bool:
    if not CACHE_PATH.exists():
        return False
    cache_mtime = CACHE_PATH.stat().st_mtime
    if Path(__file__).stat().st_mtime > cache_mtime:
        return False
    for p in _collect_source_files():
        if p.stat().st_mtime > cache_mtime:
            return False
    return True


def load_qeq_people() -> list[dict]:
    """Return the QeQ roster, rebuilding the cache on demand."""
    if _cache_is_fresh():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass  # fall through and rebuild
    try:
        return build_cache()
    except Exception as exc:
        print(f"[qeq_loader] failed to build cache: {exc}")
        return []


if __name__ == "__main__":
    entries = build_cache()
    by_country: dict[str, int] = {}
    for e in entries:
        by_country[e["country"]] = by_country.get(e["country"], 0) + 1
    print(f"QeQ cache built: {len(entries)} entries")
    for c, n in sorted(by_country.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")
    print(f"Cache file: {CACHE_PATH}")
